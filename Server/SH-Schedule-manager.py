import tkinter as tk
import tkinter.ttk as ttk
import datetime
import sqlite3 #for local databases
import pymysql #for mysql databases
import openpyxl #for excel export
import bcrypt #to hash passwords
import json #to read language data

class MainApplication(tk.Frame):
    '''a program to handle a company's weekle work schedule, the worker's data, etc.'''
    def __init__(self, parent):
        '''
        creates the main window with the main functions
        '''
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.parent.title('Beosztáskezelő')

        self.dbModule = 'sqlite3' #local database
##        self.dbModule = 'pymysql' #sql server

        with open('allTexts.txt', encoding='utf-8') as file:
            data = file.read()
        self.allTexts = json.loads(data)
        with open('helpText.txt', encoding='utf-8') as file:
            data = file.read()
        self.helpTexts = json.loads(data)
        self.language = tk.StringVar()
        self.language.set('hu')
        self.texts = self.allTexts[self.language.get()]
        
        print('Initialization...')
        self.loadDatabase('testDatabase.db') #if askopenfilename used, some error occurs
        self.updateDays()
        self.updateShifts()
        self.updateWorkers()
        print('Program ready')

        #tkinter variables to track time (actual and set year/week)
        year_week = datetime.datetime.now().isocalendar()
        #store the current time data
        self.actYear = tk.IntVar()
        self.actYear.set(year_week[0])
        self.actWeek = tk.IntVar()
        self.actWeek.set(year_week[1] + 1) #always for the next week
        #store the time data
        self.year = tk.IntVar()
        self.year.set(year_week[0])
        self.week = tk.IntVar()
        self.week.set(year_week[1] + 1) #always for the next week
        self.weekDay = tk.IntVar()
        self.weekDay.set(year_week[2])
        self.actual = self.isActual()

        #tkinter variables for widgets to change language easily
        #main (time, update, exit, tab control)
        self.yearLabelVar = tk.StringVar()
        self.weekLabelVar = tk.StringVar()
        self.updateLabelVar = tk.StringVar()
        self.exitLabelVar = tk.StringVar()
        self.settingsLabelVar = tk.StringVar()
        self.workersLabelVar = tk.StringVar()
        self.shiftsLabelVar = tk.StringVar()
        self.companyRequestsLabelVar = tk.StringVar()
        self.workerRequestsLabelVar = tk.StringVar()
        self.scheduleLabelVar = tk.StringVar()
        self.helpLabelVar = tk.StringVar()
        #settings
        self.languageLabelVar = tk.StringVar()
        self.hungarianLabelVar = tk.StringVar()
        self.englishLabelVar = tk.StringVar()
        self.saveSettingsLabelVar = tk.StringVar()
        #workers
        self.nameLabelVar = tk.StringVar()
        self.addWorkerLabelVar = tk.StringVar()
        self.deleteWorkerLabelVar = tk.StringVar()
        self.dataLabelVar = tk.StringVar()
        self.saveDataLabelVar = tk.StringVar()
        self.dateOfBirthLabelVar = tk.StringVar()
        self.phoneNumberLabelVar = tk.StringVar()
        self.membershipValidityLabelVar = tk.StringVar()
        self.activeLabelVar = tk.StringVar()
        self.passwordLabelVar = tk.StringVar()
        #shifts
        self.saveShiftsLabelVar = tk.StringVar()
        self.nameOfNewShiftLabelVar = tk.StringVar()
        self.addNewShiftLabelVar = tk.StringVar()
        #company requests
        self.saveCompanyRequestsLabelVar = tk.StringVar()
        #worker requests
        self.saveWorkerRequestsLabelVar = tk.StringVar()
        #schedule
        self.createScheduleLabelVar = tk.StringVar()
        self.fillScheduleLabelVar = tk.StringVar()
        self.algorithmLabelVar = tk.StringVar()
        self.exportToExcelLabelVar = tk.StringVar()
        #help
        self.helpTextLabelVar = tk.StringVar()

        #tkinter widgets
        #time, update, exit
        tk.Label(self.parent, textvariable=self.yearLabelVar).grid(row=0, column=0)
        tk.Entry(self.parent, textvariable=self.year, width=8).grid(row=0, column=1)
        tk.Label(self.parent, textvariable=self.weekLabelVar).grid(row=0, column=2)
        tk.Entry(self.parent, textvariable=self.week, width=8).grid(row=0, column=3)
        tk.Button(self.parent, textvariable=self.updateLabelVar, command=self.update).grid(row=0, column=4)
        tk.Button(self.parent, textvariable=self.exitLabelVar, command=self.parent.destroy).grid(row=0, column=5)

        #tab control
        self.tabControl = ttk.Notebook(self.parent)

        #settings
        #create gui
        tabSettings = ttk.Frame(self.tabControl) #tab_idx = 0
        self.tabControl.add(tabSettings, text=self.texts['Settings'])
        tk.Label(tabSettings, textvariable=self.settingsLabelVar, font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.settingsFrame = tk.Frame(tabSettings, borderwidth=2, relief='ridge')
        self.settingsFrame.grid(row=1, column=0, sticky='W')
        #update gui with self.updateSettingsGUI()
        self.updateSettingsGUI()
        
        #workers
        #create gui
        tabWorkers = ttk.Frame(self.tabControl) #tab_idx = 1
        self.tabControl.add(tabWorkers, text=self.texts['Workers'])
        tk.Label(tabWorkers, textvariable=self.workersLabelVar, font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.workerDataFrame = tk.Frame(tabWorkers, borderwidth=2, relief='ridge')
        self.workerDataFrame.grid(row=1, column=0, sticky='W')
        #update gui with self.updateWorkersGUI()
        self.updateWorkersGUI()
        #update data with self.updateWorkersData()
        #self.updateWorkersData() at the end of this block

        #shifts
        #create gui
        tabShifts = ttk.Frame(self.tabControl) #tab_idx = 2
        self.tabControl.add(tabShifts, text=self.texts['Shifts'])
        tk.Label(tabShifts, textvariable=self.shiftsLabelVar, font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.shiftManagerFrame = tk.Frame(tabShifts, borderwidth=2, relief='ridge')
        self.shiftManagerFrame.grid(row=1, column=0, sticky='W')
        #update data with self.updateShiftManagerData()
        #self.updateShiftManagerData() at the end of this block

        #company requests
        #create gui
        tabCompanyRequests = ttk.Frame(self.tabControl) #tab_idx = 3
        self.tabControl.add(tabCompanyRequests, text=self.texts['Company requests'])
        tk.Label(tabCompanyRequests, textvariable=self.companyRequestsLabelVar, font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.companyRequestFrame = tk.Frame(tabCompanyRequests, borderwidth=2, relief='ridge')
        self.companyRequestFrame.grid(row=1, column=0, sticky='W')
        #update gui with self.updateCompanyRequestsGUI()
        self.updateCompanyRequestsGUI()
        #update data with self.updateCompanyRequestsData()
        #self.updateCompanyRequestsData() at the end of this block

        #worker requests
        #create gui
        tabWorkerRequests = ttk.Frame(self.tabControl) #tab_idx = 4
        self.tabControl.add(tabWorkerRequests, text=self.texts['Worker requests'])
        tk.Label(tabWorkerRequests, textvariable=self.workerRequestsLabelVar, font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.workerRequestFrame = tk.Frame(tabWorkerRequests, borderwidth=2, relief='ridge')
        self.workerRequestFrame.grid(row=1, column=0, sticky='W')
        #update gui with self.updateWorkerRequestsGUI()
        self.updateWorkerRequestsGUI()
        #update data with self.updateWorkerRequestsData()
        #self.updateWorkerRequestsData() at the end of this block

        #schedule creation and checking
        #create gui
        tabSchedules = ttk.Frame(self.tabControl) #tab_idx = 5
        self.tabControl.add(tabSchedules, text=self.texts['Schedule'])
        tk.Label(tabSchedules, textvariable=self.scheduleLabelVar, font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.scheduleFrame = tk.Frame(tabSchedules, borderwidth=2, relief='ridge')
        self.scheduleFrame.grid(row=2, column=0, sticky='W')
        #update gui with self.updateSchedulesGUI()
        self.updateSchedulesGUI()
        #update data with self.updateSchedulesData()
        #self.updateSchedulesData() at the end of this block

        #help
        #create gui
        tabHelp = ttk.Frame(self.tabControl) #tab_idx = 6
        self.tabControl.add(tabHelp, text=self.texts['Help'])
        tk.Label(tabHelp, textvariable=self.helpLabelVar, font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.helpFrame = tk.Frame(tabHelp, borderwidth=2, relief='ridge')
        self.helpFrame.grid(row=2, column=0, sticky='W')
        #update gui with self.updateHelpGUI()
        self.updateHelpGUI()

        self.tabControl.grid(row=1, column=0, columnspan=20)

        #update widget text variables
        self.updateWidgetVariables()
        
        #update data
        #self.update()
        self.updateWorkersData()
        self.updateShiftManagerData()
        self.updateCompanyRequestsData()
        self.updateWorkerRequestsData()
        self.updateSchedulesData()

    def updateWidgetVariables(self):
        #main (time, update, exit, tab control)
        self.yearLabelVar.set(self.texts['Year'])
        self.weekLabelVar.set(self.texts['Week'])
        self.updateLabelVar.set(self.texts['Update'])
        self.exitLabelVar.set(self.texts['Exit'])
        self.tabControl.tab(0, text=self.texts['Settings'])
        self.settingsLabelVar.set(self.texts['Settings'])
        self.tabControl.tab(1, text=self.texts['Workers'])
        self.workersLabelVar.set(self.texts['Workers'])
        self.tabControl.tab(2, text=self.texts['Shifts'])
        self.shiftsLabelVar.set(self.texts['Shifts'])
        self.tabControl.tab(3, text=self.texts['Company requests'])
        self.companyRequestsLabelVar.set(self.texts['Company requests'])
        self.tabControl.tab(4, text=self.texts['Worker requests'])
        self.workerRequestsLabelVar.set(self.texts['Worker requests'])
        self.tabControl.tab(5, text=self.texts['Schedule'])
        self.scheduleLabelVar.set(self.texts['Schedule'])
        self.tabControl.tab(6, text=self.texts['Help'])
        self.helpLabelVar.set(self.texts['Help'])
        #settings
        self.languageLabelVar.set(self.texts['Language'])
        self.hungarianLabelVar.set(self.texts['Hungarian'])
        self.englishLabelVar.set(self.texts['English'])
        self.saveSettingsLabelVar.set(self.texts['Save settings'])
        #workers
        self.nameLabelVar.set(self.texts['Name'])
        self.addWorkerLabelVar.set(self.texts['Add worker'])
        self.deleteWorkerLabelVar.set(self.texts['Delete worker'])
        self.dataLabelVar.set(self.texts['Data'])
        self.saveDataLabelVar.set(self.texts['Save data'])
        self.dateOfBirthLabelVar.set(self.texts['Date of birth'])
        self.phoneNumberLabelVar.set(self.texts['Phone number'])
        self.membershipValidityLabelVar.set(self.texts['Membership validity'])
        self.activeLabelVar.set(self.texts['Active'])
        self.passwordLabelVar.set(self.texts['Password'])
        #shifts
        self.saveShiftsLabelVar.set(self.texts['Save shifts'])
        self.nameOfNewShiftLabelVar.set(self.texts['Name of new shift'])
        self.addNewShiftLabelVar.set(self.texts['Add new shift'])
        #company requests
        self.saveCompanyRequestsLabelVar.set(self.texts['Save company requests'])
        #worker requests
        self.saveWorkerRequestsLabelVar.set(self.texts['Save worker requests'])
        #schedule
        self.createScheduleLabelVar.set(self.texts['Create schedule'])
        self.fillScheduleLabelVar.set(self.texts['Fill schedule'])
        self.algorithmLabelVar.set(self.texts['Algorithm'])
        self.exportToExcelLabelVar.set(self.texts['Export to Excel'])
        #help
        self.helpTextLabelVar.set(self.helpTexts[self.language.get()])

    def isActual(self):
        if self.actYear.get() == self.year.get() and self.actWeek.get() == self.week.get():
            actual = 0 #present
        elif self.actYear.get() >= self.year.get() and self.actWeek.get() >= self.week.get():
            actual = -1 #past
        else:
            actual = 1 #future
##        print('actual:', actual)
        return actual
        
    def loadDatabase(self, dataBaseFilename=''):
        '''
        loads the database of the given name
        the open file dialog is not working
        '''
        if dataBaseFilename == '':
            self.dataBaseFilename = tk.filedialog.askopenfilename(title='Adatbázis betöltése')
        else:
            self.dataBaseFilename = dataBaseFilename

        if self.dbModule == 'sqlite3':
            self.ph = '?' #placeholder character in sql statements/queries
            self.connection = sqlite3.connect(self.dataBaseFilename)
        elif self.dbModule == 'pymysql':
            self.dataBaseFilename = 'test1'
            self.ph = '%s' #placeholder character in sql statements/queries
            self.connection = pymysql.connect(host='localhost', user='root', password='', database=self.dataBaseFilename)
        self.cursor = self.connection.cursor()
        print('Database: "' + self.dataBaseFilename + '" loaded')

    def updateDays(self):
        '''
        lists the days from the database
        '''
        self.cursor.execute('SELECT dayName FROM days ORDER BY dayId')
        arrayDays = self.cursor.fetchall()
        self.days = []
        for i in range(0, len(arrayDays)):
            self.days.append(arrayDays[i][0])
        print('updateDays done')

    def updateShifts(self):
        '''
        lists the (active?) shifts from the database
        '''
        self.shifts, self.allShifts = [], []
        #isActive feature is under development
        #only active shifts
        self.cursor.execute('SELECT shiftName FROM shifts WHERE isActive = 1 ORDER BY shiftId')
        arrayShifts = self.cursor.fetchall()
        for i in range(0, len(arrayShifts)):
            self.shifts.append(arrayShifts[i][0])
        #all shifts (for shift manager)
        self.cursor.execute('SELECT shiftName FROM shifts ORDER BY shiftId')
        arrayShifts = self.cursor.fetchall()
        for i in range(0, len(arrayShifts)):
            self.allShifts.append(arrayShifts[i][0])
        print('updateShifts done')

    def updateWorkers(self):
        '''
        lists the workers from the database sorted by name
        '''
        self.cursor.execute('SELECT workerName FROM workers')
        self.workerNames = []
        workerNamesFetchall = self.cursor.fetchall()
        if workerNamesFetchall != []:
            for row in workerNamesFetchall:
                self.workerNames.append(row[0])
        else:
            self.workerNames.append('')
        self.workerNames.sort()
        print('updateWorkers done')

    def update(self):
        '''
        update all
        '''
        print('Updating...')
        try:
            self.year.get()
            self.week.get()
        except:
            self.year.set(self.actYear.get())
            self.week.set(self.actWeek.get())
        self.actual = self.isActual()

        self.updateShifts()
        #update gui
        self.updateCompanyRequestsGUI()
        self.updateWorkerRequestsGUI()
        self.updateSchedulesGUI()
        #update data - always
        self.updateWorkersData()
        self.updateShiftManagerData()

        #update data - only for the actual year and week
        if self.isActual:
            self.updateCompanyRequestsData()
            self.updateWorkerRequestsData()
            #self.updateSchedulesData() #update only here, when previous schedule are not available
        else:
            pass
        self.updateSchedulesData()

        print('...update done')

    def quit(self):
        '''
        saves the database and closes the program
        '''
        print('Closing...')
        self.saveDatabase()
        self.connection.close()
        self.mainWindow.destroy()

    def saveDatabase(self):
        '''
        saves the database
        '''
        self.connection.commit()
        print('Database saved')


#------------------------------------------------------------------------------------------------------
#Settings

    def updateSettingsGUI(self):
        tk.Label(self.settingsFrame, textvariable=self.languageLabelVar, font=('Helvetica 10 bold')).grid(row=0, column=0)
        tk.Radiobutton(self.settingsFrame, textvariable=self.hungarianLabelVar, variable=self.language, value='hu').grid(row=0, column=1)
        tk.Radiobutton(self.settingsFrame, textvariable=self.englishLabelVar, variable=self.language, value='en').grid(row=0, column=2)
        tk.Button(self.settingsFrame, textvariable=self.saveSettingsLabelVar, command=self.saveSettings).grid(row=1, column=0)

    def saveSettings(self):
        self.texts = self.allTexts[self.language.get()]
        self.updateWidgetVariables()
        

#------------------------------------------------------------------------------------------------------
#Worker data

    def updateWorkersGUI(self):
        tk.Label(self.workerDataFrame, textvariable=self.nameLabelVar).grid(row=0, column=0)
        self.workerName = tk.StringVar() #workerName variable
        self.workerName.set('')
        self.nameOptionsW = ttk.Combobox(self.workerDataFrame, width=18, textvariable=self.workerName, height=5) #why combobox, why not optionmenu? -> to write new names!
        self.nameOptionsW.bind('<<ComboboxSelected>>', self.nameOptionsWSelectionEvent)
        self.nameOptionsW.grid(row=0, column=1)
        tk.Button(self.workerDataFrame, textvariable=self.addWorkerLabelVar, command=self.addWorker).grid(row=0, column=2)
        tk.Button(self.workerDataFrame, textvariable=self.deleteWorkerLabelVar, command=self.deleteWorker).grid(row=0, column=3)
        tk.Label(self.workerDataFrame, textvariable=self.dataLabelVar, font=('Helvetica 10 bold')).grid(row=2, column=0, columnspan=2, sticky='W')
        tk.Button(self.workerDataFrame, textvariable=self.saveDataLabelVar, command=self.saveWorkerData).grid(row=3, column=2)
        tk.Label(self.workerDataFrame, textvariable=self.dateOfBirthLabelVar).grid(row=3, column=0)
        self.dateOfBirthVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.dateOfBirthVariable).grid(row=3, column=1)
        tk.Label(self.workerDataFrame, textvariable=self.phoneNumberLabelVar).grid(row=4, column=0)
        self.phoneNumberVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.phoneNumberVariable).grid(row=4, column=1)
        tk.Label(self.workerDataFrame, textvariable=self.membershipValidityLabelVar).grid(row=5, column=0)
        self.membershipValidityVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.membershipValidityVariable).grid(row=5, column=1)
        tk.Label(self.workerDataFrame, textvariable=self.activeLabelVar).grid(row=6, column=0)
        self.isActiveVariable = tk.BooleanVar()
        self.isActiveCheckbutton = tk.Checkbutton(self.workerDataFrame, variable=self.isActiveVariable)
        self.isActiveCheckbutton.grid(row=6, column=1)
        tk.Label(self.workerDataFrame, textvariable=self.passwordLabelVar).grid(row=7, column=0)
        self.passwordVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.passwordVariable, show='*').grid(row=7, column=1)
        print('updateWorkersGUI done')

    def updateWorkersData(self):
        '''
        load actual data from database and update gui
        '''
        self.updateWorkers() #list of workers
        self.nameOptionsW['values'] = self.workerNames
        self.nameOptionsWR['values'] = self.workerNames

    def nameOptionsWSelectionEvent(self, event):
        '''
        this function is called when you select a name from the dropdown list
        it loads the data of the selected worker
        '''
        workerName = self.workerName.get()
        self.cursor.execute('SELECT dateOfBirth FROM workers WHERE workerName = ' + self.ph, (workerName, ))
        self.dateOfBirthVariable.set( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT phoneNumber FROM workers WHERE workerName = ' + self.ph, (workerName, ))
        self.phoneNumberVariable.set( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT membershipValidity FROM workers WHERE workerName = ' + self.ph, (workerName, ))
        self.membershipValidityVariable.set( self.cursor.fetchone()[0] )
        #isActive feature is not working yet
        #self.isActiveVariable.set( True if self.cursor.fetchone()[0] == 1 else False )
        self.cursor.execute('SELECT isActive FROM workers WHERE workerName = ' + self.ph, (workerName, ))
        if self.cursor.fetchone()[0] == 1:
            self.isActiveCheckbutton.select()
        else:
            self.isActiveCheckbutton.deselect()
        self.passwordVariable.set('')

    def addWorker(self):
        '''
        adds the worker with the given name to the database
        calls saveWorkerData() to save the other data for the worker
        '''
        workerName = self.workerName.get()
        if workerName not in ['', 'név', 'name']:
            self.saveWorkerData()
            self.updateWorkersData()
            print(workerName + ' added')

    def saveWorkerData(self):
        '''
        saves data (birthday, phone number, etc.) for the worker
        '''
        workerName = self.workerName.get()
        dateOfBirth = self.dateOfBirthVariable.get()
        phoneNumber = self.phoneNumberVariable.get()
        membershipValidity = self.membershipValidityVariable.get()
        isActive = self.isActiveVariable.get()
        password = self.passwordVariable.get()
        hashedPassword = bcrypt.hashpw(bytes(password, "utf-8"), bcrypt.gensalt()) #hash the password and store this value
        self.cursor.execute('SELECT 1 FROM workers WHERE workerName = ' + self.ph, (workerName, )) #check if record already exists in the database
        exists = self.cursor.fetchone() #None or 1
        #print('worker exists:', exists)
        if not exists: #if the worker is not in the database, insert
            self.cursor.execute('INSERT INTO workers (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive, password) VALUES (' + self.ph + ', ' + self.ph + ', ' + self.ph + ', ' + self.ph + ', ' + self.ph + ', ' + self.ph + ')',
                                (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive, hashedPassword))
        else: #if the worker is already in the database, update
            self.cursor.execute('UPDATE workers SET dateOfBirth = ' + self.ph + ' WHERE workerName = ' + self.ph, (dateOfBirth, workerName))
            self.cursor.execute('UPDATE workers SET phoneNumber = ' + self.ph + '  WHERE workerName = ' + self.ph, (phoneNumber, workerName))
            self.cursor.execute('UPDATE workers SET membershipValidity = ' + self.ph + '  WHERE workerName = ' + self.ph, (membershipValidity, workerName))
            self.cursor.execute('UPDATE workers SET isActive = ' + self.ph + ' WHERE workerName = ' + self.ph, (str(int(isActive)), workerName))
            self.cursor.execute('UPDATE workers SET password = ' + self.ph + ' WHERE workerName = ' + self.ph, (hashedPassword, workerName))
        self.saveDatabase()
        print(workerName + ' data saved')
        
    def deleteWorker(self):
        '''
        deletes the worker with the given name from the database
        '''
        workerName = self.workerName.get()
        self.cursor.execute('DELETE FROM workers WHERE workerName = ' + self.ph, (workerName, ))
        self.saveDatabase()
        self.updateWorkersData()
        print(workerName + ' deleted')


#------------------------------------------------------------------------------------------------------
#Shift manager

    def updateShiftManagerData(self):
        for child in self.shiftManagerFrame.winfo_children():
            child.destroy()
        tk.Button(self.shiftManagerFrame, textvariable=self.saveShiftsLabelVar, command=self.saveShifts).grid(row=len(self.allShifts)+2, column=0)
        tk.Label(self.shiftManagerFrame, textvariable=self.nameOfNewShiftLabelVar).grid(row=1, column=0)
        self.newShiftName = tk.StringVar()
        tk.Entry(self.shiftManagerFrame, textvariable=self.newShiftName).grid(row=1, column=1)
        tk.Button(self.shiftManagerFrame, textvariable=self.addNewShiftLabelVar, command=self.addShift).grid(row=1, column=2)
        self.shiftCheckbuttons, self.shiftVariables = [], []
        for i in range(0, len(self.allShifts)):
            tk.Label(self.shiftManagerFrame, text=self.allShifts[i], width=8).grid(row=2+i, column=0)
            self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ' + self.ph, (self.allShifts[i], ))
            isActive = self.cursor.fetchone()[0]
            variable = tk.BooleanVar()
            variable.set(isActive)
            checkbutton = tk.Checkbutton(self.shiftManagerFrame, variable=variable)
            checkbutton.grid(row=2+i, column=1)
            self.shiftCheckbuttons.append(checkbutton)
            self.shiftVariables.append(variable)
        print('Shift manager data updated')

    def addShift(self):
        newShiftName = self.newShiftName.get()
        if newShiftName != '':
            self.cursor.execute('INSERT INTO shifts (shiftName, isActive) VALUES (' + self.ph + ',' + self.ph + ')', (newShiftName, 0, )) #set as active!
        #add 0 as workerNumber for companyRequests, else error in loading
        self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (newShiftName, ))
        shiftId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('INSERT INTO companyRequests (dayId, shiftId, workerNumber) VALUES (' + self.ph + ',' + self.ph + ',' + self.ph + ')', (j, shiftId, 0))
        self.saveDatabase()
        #update shifts and shift manager
        self.updateShifts()
        self.updateShiftManagerData()
        print(newShiftName + ' shift added')

    def saveShifts(self):
        for i in range(0, len(self.allShifts)):
            shiftName = self.allShifts[i]
            isActive = self.shiftVariables[i].get()
            isActive = 1 if isActive == True else 0
            self.cursor.execute('UPDATE shifts SET isActive = ' + self.ph + ' WHERE shiftName = ' + self.ph, (str(isActive), shiftName))
        self.saveDatabase()
        print('Shifts saved')


#------------------------------------------------------------------------------------------------------
#Company requests

    def updateCompanyRequestsGUI(self):
        for child in self.companyRequestFrame.winfo_children():
            child.destroy()
        tk.Button(self.companyRequestFrame, textvariable=self.saveCompanyRequestsLabelVar, command=self.saveCompanyRequest).grid(row=1, column=1, columnspan=2)
        #create the field of entries
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        date = datetime.datetime.fromisocalendar(year, week, weekDay)
        startDate = date - datetime.timedelta(days=self.weekDay.get()-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            tk.Label(self.companyRequestFrame, text=text, width=8).grid(row=2, column=1+j)
        for i in range(0, len(self.shifts)):
            tk.Label(self.companyRequestFrame, text=self.shifts[i], width=8).grid(row=3+i, column=0)
        self.companyRequestEntries, self.companyRequestVariables = [], [] #lists to store the entries and their variables
        for j in range(0, len(self.days)):
            self.companyRequestEntries.append([])
            self.companyRequestVariables.append([])
            for i in range(0, len(self.shifts)):
                variable = tk.IntVar()
                variable.set(-1)
                entry = tk.Entry(self.companyRequestFrame, textvariable=variable, width=5)
                if self.actual == -1: #if past, not present or future
                    entry['state'] = 'disabled'
                entry.grid(row=3+i, column=1+j)
                self.companyRequestEntries[j].append(entry)
                self.companyRequestVariables[j].append(variable)
        print('updateCompanyRequestsGUI done')

    def updateCompanyRequestsData(self):
        '''
        loads company requests for the given week
        and fills the previousley created entry table with the data
        '''
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerNumber FROM companyRequests WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) )
                workerNumber = self.cursor.fetchone()[0]
                self.companyRequestVariables[j][i].set(workerNumber)
        print('Company requests data updated')

    def getCompanyRequest(self):
        '''
        takes the numbers from the entry table into a numpy array
        '''
        self.companyRequestGrid = [[0 for j in range(len(self.days))] for i in range(len(self.shifts))]
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.companyRequestGrid[i][j] = self.companyRequestVariables[j][i].get()
        #print(self.companyRequestGrid)

    def saveCompanyRequest(self):
        '''
        saves company requests to the database
        first calls getCompanyRequest() in order to get the data from the entry field
        '''
        self.getCompanyRequest()
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
##                self.cursor.execute('INSERT OR IGNORE INTO companyRequest (dayID, shiftId, workerNumber) VALUES (?, ?, ?)',
##                                    (dayId, shiftId, int(self.companyRequestGrid[i][j])) ) #cast a numpy value to int: value.item()
                self.cursor.execute('UPDATE companyRequests SET workerNumber = ' + self.ph +
                                    ' WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, ( str(int(self.companyRequestGrid[i][j])), str(dayId), str(shiftId) ) )
                #update és insert egyszerre: ha a meglévő érték nem azonos a beírttal, frissíteni kell
        self.saveDatabase()
        print('Company requests saved')

#------------------------------------------------------------------------------------------------------
#Company requests - Shift manager

        
#------------------------------------------------------------------------------------------------------
#Worker requests

    def updateWorkerRequestsGUI(self):
        for child in self.workerRequestFrame.winfo_children():
            child.destroy()
        self.nameOptionsWR = ttk.Combobox(self.workerRequestFrame, width=20, textvariable=self.workerName)
        self.nameOptionsWR.bind('<<ComboboxSelected>>', self.nameOptionsWRSelectionEvent)
        self.nameOptionsWR.grid(row=0, column=0, columnspan=3)
        tk.Button(self.workerRequestFrame, textvariable=self.saveWorkerRequestsLabelVar, command=self.saveWorkerRequest).grid(row=0, column=3)
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        date = datetime.datetime.fromisocalendar(year, week, weekDay)
        startDate = date - datetime.timedelta(days=self.weekDay.get()-1)
        #endDate = startDate + datetime.timedelta(days=6)
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            tk.Label(self.workerRequestFrame, text=text, width=8).grid(row=1, column=1+j)
        for i in range(0, len(self.shifts)):
            tk.Label(self.workerRequestFrame, text=self.shifts[i], width=8).grid(row=2+i, column=0)
##        self.requestCheckbuttons, self.requestVariables = [], [] #lists to store the entries and their variables
        self.requestCheckbuttons, self.requestVariables = {}, {} #dictionaries to store the entries and their variables (to label shifts)
        for j in range(0, len(self.days)):
##            self.requestCheckbuttons.append([])
##            self.requestVariables.append([])
            self.requestCheckbuttons[j] = {}
            self.requestVariables[j] = {}
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerNumber FROM companyRequests WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) ) #str(dayId) may not work!
                #a checkbutton should be active only if
                #the requested number of workers is greater than 0 for the given shift
                variable = tk.BooleanVar()
                checkbutton = tk.Checkbutton(self.workerRequestFrame, variable=variable)
                checkbutton.grid(row=2+i, column=1+j)
                if self.cursor.fetchone()[0] > 0 and self.actual >= 0: #if self.actual = 0 or 1 (present or future)
                    #if workers needed and the time is actual
                    checkbutton['state'] = 'normal'
                else:
                    #in all other cases can't be set
                    checkbutton['state'] = 'disabled'
##                self.requestCheckbuttons[j].append(checkbutton)
##                self.requestVariables[j].append(variable)
                self.requestCheckbuttons[j][shiftId] = checkbutton
                self.requestVariables[j][shiftId] = variable
        print('updateWorkerRequestsGUI done')

    def updateWorkerRequestsData(self):
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS workerRequests_' + str(year) + '_' + str(week) + 
                            '(workerId INT, dayId INT, shiftId INT, UNIQUE(workerId, dayId, shiftId))')
        print('Worker request data updated')

    def nameOptionsWRSelectionEvent(self, event):
        '''
        event for selecting a name
        first it deselects all checkbuttons
        then it checks the shifts the worker requested for the given week
        '''
        for daysCheckbuttons in self.requestCheckbuttons.values():
            for checkbutton in daysCheckbuttons.values():
                checkbutton.deselect()
##        for daysCheckbuttons in self.requestCheckbuttons:
##            for checkbutton in daysCheckbuttons:
##                checkbutton.deselect()
        year = self.year.get()
        week = self.week.get()
        workerName = self.workerName.get()
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ' + self.ph, (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                try:
                    self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) ) #str(dayId) did not work!
                    workerIds = [row[0] for row in self.cursor.fetchall()]
                    if workerId in workerIds:
                        self.requestCheckbuttons[j][shiftId].select()
                except:
                    pass

    def getWorkerRequest(self):
        '''
        takes the checks from the check table into a numpy array (1 if checked, else 0)
        '''
        workerName = self.workerName.get()
        
        self.workerRequestGrid = {}
        for i in range(0, len(self.shifts)):
            self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
            shiftId = self.cursor.fetchone()[0]
            self.workerRequestGrid[shiftId] = {}
            for j in range(len(self.days)):
                self.workerRequestGrid[shiftId][j] = 1 if self.requestVariables[j][shiftId].get() else 0
                
##        self.workerRequestGrid = [[0 for j in range(len(self.days))] for i in range(len(self.shifts))]
##        for j in range(0, len(self.days)):
##            for i in range(0, len(self.shifts)):
##                self.workerRequestGrid[i][j] = 1 if self.requestVariables[j][i].get() else 0 #when creating these checkbuttons and variables, the indices are reversed
        #print(workerName, '\n', self.workerRequestGrid)

    def saveWorkerRequest(self):
        '''
        saves worker requests for the given week to the database 
        '''
        self.getWorkerRequest()
        workerName = self.workerName.get()
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS workerRequests_' + str(year) + '_' + str(week) + 
                            '(workerId, dayId, shiftId, UNIQUE(workerId, dayId, shiftId))')
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ' + self.ph, (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerNumber FROM companyRequests WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) )
                if self.workerRequestGrid[shiftId][j] == 1:
                    self.cursor.execute('INSERT OR IGNORE INTO workerRequests_' + str(year) + '_' + str(week) +
                                        ' (workerId, dayId, shiftId) VALUES (' + self.ph + ',' + self.ph + ',' + self.ph + ')', (workerId, dayId, shiftId))
##                if self.workerRequestGrid[i][j] == 1:
##                    self.cursor.execute('INSERT OR IGNORE INTO workerRequests_' + str(year) + '_' + str(week) +
##                                        ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', (workerId, j, i))
                else:
                    try:
                        self.cursor.execute('DELETE FROM workerRequests_' + str(year) + '_' + str(week) +
                                            ' WHERE workerId = ' + self.ph + ' AND dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (workerId, dayId, shiftId) )
                    except:
                        pass
        self.saveDatabase()
        print('Worker request saved for', workerName)

#------------------------------------------------------------------------------------------------------
#Schedule manager

    def updateSchedulesGUI(self):
        for child in self.scheduleFrame.winfo_children():
            child.destroy()
        tk.Button(self.scheduleFrame, textvariable=self.createScheduleLabelVar, command=self.createSchedule).grid(row=0, column=0, columnspan=2)
        tk.Button(self.scheduleFrame, textvariable=self.fillScheduleLabelVar, command=self.fillSchedule).grid(row=0, column=2, columnspan=2)
        tk.Label(self.scheduleFrame, textvariable=self.algorithmLabelVar).grid(row=0, column=4)
        self.algorithmList = ['random', 'frommin']
        self.algorithmVar = tk.StringVar()
        self.algorithmVar.set(self.algorithmList[0])
        tk.OptionMenu(self.scheduleFrame, self.algorithmVar, *self.algorithmList).grid(row=0, column=5)
        tk.Button(self.scheduleFrame, textvariable=self.exportToExcelLabelVar, command=self.scheduleExportXlsx).grid(row=0, column=6, columnspan=2)
        self.parent.bind('<Enter>', lambda event: self.highlightOn(event, frame=self.scheduleFrame))
        self.parent.bind('<Leave>', lambda event: self.highlightOff(event, frame=self.scheduleFrame))
        self.scheduleByHandCheckbuttons, self.scheduleByHandVariables, self.scheduleByHandNameLabels = [], [], []
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        tk.Label(self.scheduleFrame, text=str(year)+'/'+str(week)).grid(row=1, column=0)
        date = datetime.datetime.fromisocalendar(year, week, weekDay)
        startDate = date - datetime.timedelta(days=self.weekDay.get()-1)
        #endDate = startDate + datetime.timedelta(days=6)
        row = 1
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            tk.Label(self.scheduleFrame, text=text, width=12, font='Helvetica 10 bold').grid(row=row, column=1+2*j, columnspan=2) #!!!!!!!!! column(span)
        row += 1
        if self.actual == 0: #if date is actual
            maxWorkerNumbers = self.getMaxWorkerNumbers('workerRequests') #determine max requests for the shifts (for drawing)
            for i in range(0, len(self.shifts)):
                tk.Label(self.scheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=row, column=0)
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                row = row + maxWorkerNumbers[shiftId]
        elif self.actual == -1: #if date is in the past
            try:
                maxWorkerNumbers = self.getMaxWorkerNumbers('schedule') #determine max requests for the shifts (for drawing)
                #print('maxWorkerNumbers:', maxWorkerNumbers)
                self.cursor.execute('SELECT shiftId FROM schedule_' + str(year) + '_' + str(week))
                shiftIds = [id_[0] for id_ in self.cursor.fetchall()]
                shiftIds = list(set(shiftIds)) #to get unique elements
                shiftIds.sort()
                shifts = [ self.allShifts[k] for k in shiftIds]
                #print('shifts:', shifts)
                for i in range(0, len(shifts)):
                    tk.Label(self.scheduleFrame, text=shifts[i], width=8, font='Helvetica 10 bold').grid(row=row, column=0)
                    row = row + maxWorkerNumbers[i]
            except:
                print('Table does not exist')
        else: #if date is in the future
            print('Date is in the future')
        print('updateSchedulesGUI done')

    def scheduleExportXlsx(self):
        '''
        exports the schedule for the given week into a .xlsx file
##        first loads the schedule from the database
##        saves the backup workers for the week on a different worksheet (same as loading the scheduled workers)
        '''
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        filename = 'schedule_' + str(year) + '_' + str(week) + '.xlsx'
        #schedule
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'schedule_' + str(year) + '_' + str(week)
        #requests = [4, 1, 4], better solution below
        requests = self.getMaxWorkerNumbers('schedule') #gives the max number of requests for shifts
        print('requests:', requests)
        row = 2
        worksheet.cell(row=1, column=1).value = str(year) + '/' + str(week)
        worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        date = datetime.datetime.fromisocalendar(year, week, weekDay)
        startDate = date - datetime.timedelta(days=self.weekDay.get()-1)
        #endDate = startDate + datetime.timedelta(days=6)
        
        for j in range(0, len(self.days)):
            dayDate = startDate + datetime.timedelta(days=j)
            month, day = dayDate.month, dayDate.day
            text = str(month) + '.' + str(day) + '.\n' + self.days[j]
            #text = self.days[j]
            worksheet.cell(row=1, column=2+j).value = text
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
            
        self.cursor.execute('SELECT shiftId FROM schedule_' + str(year) + '_' + str(week))
        shiftIds = [id_[0] for id_ in self.cursor.fetchall()]
        shiftIds = list(set(shiftIds)) #to get unique elements
        shiftIds.sort()
        shifts = [ self.allShifts[k] for k in shiftIds]
        print('shifts:', shifts)
        for i in range(0, len(shifts)):
            shiftName = shifts[i]
            self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (shiftName, ))
            shiftId = self.cursor.fetchone()[0]
            worksheet.cell(row=row, column=1).value = shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[shiftId]
        
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                for k in range(0, requests[shiftId]):
                    try:
                        workerId = self.schedule[dayId][shiftId][k]
                        self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ' + self.ph, (workerId, ))
                        workerName = self.cursor.fetchone()[0]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[shiftId]
        
        workbook.save(filename=filename)
        print('Schedule exported')

    def fillSchedule(self):
        pass
    
    def createSchedule(self):
        if self.actual == 0: #if date is actual (or future?)
            '''
            creates schedule from the check table
            also calls createBackup() #not yet
            '''
            year = self.year.get()
            week = self.week.get()
            self.cursor.execute('DROP TABLE IF EXISTS schedule_'  + str(year) + '_' + str(week))
            self.cursor.execute('CREATE TABLE schedule_'  + str(year) + '_' + str(week) +
                                '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
            for day in range(0, len(self.scheduleByHandVariables)):
                for shift in range(0, len(self.scheduleByHandVariables[day])):
                    for row in self.scheduleByHandVariables[day][shift]:
                        if row[0].get() == True:
                            self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                                '(workerId, dayId, shiftId) VALUES (' + self.ph + ',' + self.ph + ',' + self.ph +')', (row[1], day, shift) )
##            self.createBackup() #not yet
            self.saveDatabase()
            print('Schedule created')
        else:
            print('Invalid date')

    def getMaxWorkerNumbers(self, table):
        #should return a dictionary: { shiftId_1 : number_1, ... }
        maxWorkerNumbers = {}
##        maxWorkerNumbers = [1]*len(self.shifts)
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS workerRequests_' + str(year) + '_' + str(week) + 
                            ' (workerId INT, dayId INT, shiftId INT, UNIQUE(workerId, dayId, shiftId))')
        if table == 'workerRequests':
            for j in range(0, len(self.days)):
                self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
                dayId = self.cursor.fetchone()[0]
                for i in range(0, len(self.shifts)):
                    self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
                    shiftId = self.cursor.fetchone()[0]
                    self.cursor.execute('SELECT workerId FROM ' + table + '_' + str(year) + '_' + str(week) +
                                                ' WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) )
                    workerIds = self.cursor.fetchall()
                    if shiftId not in maxWorkerNumbers.keys():
                        maxWorkerNumbers[shiftId] = 1
                    if len(workerIds) >= maxWorkerNumbers[shiftId]:
                        maxWorkerNumbers[shiftId] = len(workerIds)
##                    if len(workerIds) >= maxWorkerNumbers[i]:
##                        maxWorkerNumbers[i] = len(workerIds)
        elif table == 'schedule':
            self.cursor.execute('SELECT shiftId FROM schedule_' + str(year) + '_' + str(week))
            shiftIds = [id_[0] for id_ in self.cursor.fetchall()]
            shiftIds = list(set(shiftIds)) #to get unique elements
            shiftIds.sort()
            shifts = [ self.allShifts[k] for k in shiftIds]
            for j in range(0, len(self.days)):
                self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
                dayId = self.cursor.fetchone()[0]
                for i in range(0, len(shifts)): #not a predefined list of shifts, but the available shifts from the schedule
                    self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (shifts[i], ))
                    shiftId = self.cursor.fetchone()[0]
                    self.cursor.execute('SELECT workerId FROM ' + table + '_' + str(year) + '_' + str(week) +
                                                ' WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) )
                    workerIds = self.cursor.fetchall()
                    if shiftId not in maxWorkerNumbers.keys():
                        maxWorkerNumbers[shiftId] = 1
                    if len(workerIds) >= maxWorkerNumbers[shiftId]:
                        maxWorkerNumbers[shiftId] = len(workerIds)
##                    if len(workerIds) >= maxWorkerNumbers[i]:
##                        maxWorkerNumbers[i] = len(workerIds)
        return maxWorkerNumbers

    def highlightOn(self, event, frame):
        '''
        when the mouse hovers over a name, highlights all of his/her requests for the week in red
        '''
        #print('highlightOn')
        try:
            eventWidget = event.widget
            eventText = eventWidget['text']
            widgetList = frame.winfo_children()
            highlightList = []
            for widget in widgetList:
                if isinstance(widget, tk.Label):
                    text = widget['text']
                    if text == eventText:
                        highlightList.append(widget)
            for widget in highlightList:
                widget.configure(fg='red')
        except:
            pass

    def highlightOff(self, event, frame):
        '''
        disables highlighting defined in highlightOn()
        '''
        #print('highlightOff')
        try:
            eventWidget = event.widget
            eventText = eventWidget['text']
            widgetList = frame.winfo_children()
            highlightList = []
            for widget in widgetList:
                if isinstance(widget, tk.Label):
                    text = widget['text']
                    if text == eventText:
                        highlightList.append(widget)
            for widget in highlightList:
                widget.configure(fg='black')
        except:
            pass

    def disableWorkerSelection(self, column, row, row_k, nameToDisable):
        '''
        if someone is scheduled to work in a shfit, he/she can't work in another shift on the given day
        the possibility to check him/her into another shift is disabled
        '''
        #print(column, row, row_k, nameToDisable)
        #print(self.scheduleByHandVariables[column][row][row_k][0].get(), self.scheduleByHandNameLabels[column][row][row_k]['text'])
        if self.scheduleByHandVariables[column][row][row_k][0].get() == True:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k]['text'] == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k]['state'] = 'disabled'
        else:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k]['text'] == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k]['state'] = 'normal'

        self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[column], ))
        dayId = self.cursor.fetchone()[0]
        self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[row], ))
        shiftId = self.cursor.fetchone()[0]
        self.disableWorkerSelectionForShift(dayId, shiftId, column, row)
    
    def disableWorkerSelectionForShift(self, dayId, shiftId, column, row):
        '''
        if the workers requested by the company for a given shift is met,
        the possibility to check other workers for that shift is disabled
        '''
        year = self.year.get()
        week = self.week.get()
        requests = self.getMaxWorkerNumbers('workerRequests') #gives the max number of requests for shifts
        workersScheduledForShift = []
        workerNumberScheduled = 0
        workersScheduledForDay = []
        for k in range(0, requests[row]):
            #try-except is not the most elegant solution
            #it is for overcoming that requests list contains the max number of requests for shifts (for example [8, 1, 5])
            #and the real requests for a given day can be fewer (for example [6, 1, 4])
            #so the index k may result in out of range error
            #and company requests is [4, 1, 4]
            try:
                if self.scheduleByHandVariables[column][row][k][0].get() == True:
                    workerNumberScheduled += 1
                    workersScheduledForShift.append(self.scheduleByHandNameLabels[column][row][k]['text'])
            except:
                pass
        #print('workersScheduledForShift: ', workersScheduledForShift)

        for row_ in range(0, len(requests)):
            for k in range(0, requests[row_]):
                try:
                    if self.scheduleByHandVariables[column][row_][k][0].get() == True:
                        workersScheduledForDay.append(self.scheduleByHandNameLabels[column][row_][k]['text'])
                except:
                    pass
        #print('workersScheduledForDay: ', workersScheduledForDay)

        self.cursor.execute('SELECT workerNumber FROM companyRequests WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) )
        workerNumber = self.cursor.fetchone()[0]
        if workerNumberScheduled == workerNumber:
            try:
                for k in range(0, requests[row]):
                    name = self.scheduleByHandNameLabels[column][row][k]['text']
                    if name not in workersScheduledForShift and name not in workersScheduledForDay:
                        self.scheduleByHandCheckbuttons[column][row][k]['state'] = 'disabled'
            except:
                pass
        else:
            try:
                for k in range(0, requests[row]):
                    name = self.scheduleByHandNameLabels[column][row][k]['text']
                    if name not in workersScheduledForShift and name not in workersScheduledForDay:
                        self.scheduleByHandCheckbuttons[column][row][k]['state'] = 'normal'
            except:
                pass

    def updateSchedulesData(self):
        year = self.year.get()
        week = self.week.get()
        weekDay = self.weekDay.get()
        self.scheduleByHandCheckbuttons, self.scheduleByHandVariables, self.scheduleByHandNameLabels = [], [], []
        self.schedule = {}
        if self.actual == 0: # if actual, load from worker requests to create schedule
            maxWorkerNumbers = self.getMaxWorkerNumbers('workerRequests')
            for j in range(0, len(self.days)):
                self.scheduleByHandCheckbuttons.append([])
                self.scheduleByHandVariables.append([])
                self.scheduleByHandNameLabels.append([])
                gridRow = 2 #same as row
                gridRow_ = 2 #to track the last empty row
                self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
                dayId = self.cursor.fetchone()[0]
                for i in range(0, len(self.shifts)):
                    self.scheduleByHandCheckbuttons[j].append([])
                    self.scheduleByHandVariables[j].append([])
                    self.scheduleByHandNameLabels[j].append([])
                    self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (self.shifts[i], ))
                    shiftId = self.cursor.fetchone()[0]
                    self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) )
                    workerIds = self.cursor.fetchall()
                    for k in range(0, maxWorkerNumbers[shiftId]):
##                    for k in range(0, maxWorkerNumbers[i]):
                        try:
                            workerId = workerIds[k][0]
                            self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ' + self.ph, (str(workerId), ) )
                            workerName = self.cursor.fetchone()[0]
                            nameLabel = tk.Label(self.scheduleFrame, text=workerName)
                            nameLabel.grid(row=gridRow_, column=1+2*j) #!!!!!!!!! column
                            self.scheduleByHandNameLabels[j][i].append(nameLabel)
                            variable = tk.BooleanVar()
                            checkbutton = tk.Checkbutton(self.scheduleFrame, variable=variable, command=lambda x1=j, x2=i, x3=k, x4=workerName: self.disableWorkerSelection(x1, x2, x3, x4))
                            checkbutton.grid(row=gridRow_, column=1+2*j+1) #!!!!!!!!! column (columnspan=2 for day names?)
                            try:
                                #check if the worker to be shown is already scheduled there (in a previous run of the program)
                                self.cursor.execute( 'SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                                     ' WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) )
                                if workerId in [ workerIds[0] for workerIds in self.cursor.fetchall()]:
                                    #if a worker is scheduled, check the box
                                    checkbutton.select()
                            except:
                                pass
                            self.scheduleByHandCheckbuttons[j][i].append(checkbutton)
                            self.scheduleByHandVariables[j][i].append([variable, workerId, workerName])
                        except:
                            #shitty solution to fill empty spaces (rowconfigure?)
                            tk.Label(self.scheduleFrame, text='').grid(row=gridRow_, column=1+2*j)
                        gridRow_ += 1
                    gridRow = gridRow + maxWorkerNumbers[shiftId]
##                    gridRow = gridRow + maxWorkerNumbers[i]
        elif self.actual == -1: #if date is in the past, load from previously created schedule
            try:
                self.cursor.execute('SELECT 1 FROM schedule_' + str(year) + '_' + str(week))
                exists = 1
            except:
                exists = 0
            print('exists:', exists)
            if exists: #if schedule table exists for the week
                maxWorkerNumbers = self.getMaxWorkerNumbers('schedule')
                self.cursor.execute('SELECT shiftId FROM schedule_' + str(year) + '_' + str(week))
                shiftIds = [id_[0] for id_ in self.cursor.fetchall()]
                shiftIds = list(set(shiftIds)) #to get unique elements
                shiftIds.sort()
                shifts = [ self.allShifts[k] for k in shiftIds]
                for j in range(0, len(self.days)):
                    row = 2
                    row_ = 2 #to track the last empty row
                    self.cursor.execute('SELECT dayId FROM days WHERE dayName = ' + self.ph, (self.days[j], ))
                    dayId = self.cursor.fetchone()[0]
                    self.schedule[dayId] = {}
                    for i in range(0, len(shifts)):
                        self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ' + self.ph, (shifts[i], ))
                        shiftId = self.cursor.fetchone()[0]
                        self.schedule[dayId][shiftId] = []
                        self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + self.ph + ' AND shiftId = ' + self.ph, (dayId, shiftId) )
                        workerIds = self.cursor.fetchall()
                        for k in range(0, maxWorkerNumbers[shiftId]):
##                        for k in range(0, maxWorkerNumbers[i]):
                            try:
                                workerId = workerIds[k][0]
                                self.schedule[dayId][shiftId].append(workerId)
                                self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ' + self.ph, (str(workerId), ) )
                                workerName = self.cursor.fetchone()[0]
                                nameLabel = tk.Label(self.scheduleFrame, text=workerName)
                                nameLabel.grid(row=row_, column=1+2*j) #!!!!!!!!! column (columnspan=2 for day names?)
                            except:
                                #shitty solution to fill empty spaces (rowconfigure?)
                                tk.Label(self.scheduleFrame, text='').grid(row=row_, column=1+2*j)
                            row_ += 1
                        row = row + maxWorkerNumbers[shiftId]
##                        row = row + maxWorkerNumbers[i]
##                print(self.schedule)
            else:
                print('Table does not exist')
        else: #if actual = 1, date is in the future
            pass
        print('Schedules data updated')

#------------------------------------------------------------------------------------------------------
#Help

    def updateHelpGUI(self):
        tk.Label(self.helpFrame, textvariable=self.helpTextLabelVar, justify='left').grid(row=1, column=0)

    
        
if __name__ == "__main__":
    root = tk.Tk()
    MainApplication(root).grid(row=0, column=0)
    root.mainloop()
